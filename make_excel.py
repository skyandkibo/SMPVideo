import json
import os
import pandas as pd
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment

# 定义一个函数，用于添加信息
def add_info(data_store, vid=None, pid=None, uid=None, info=None):
    key = (vid, pid, uid)
    if key not in data_store:
        data_store[key] = info
    elif key in data_store: 
        for k,v in info.items():
            if k in data_store[key]: 
                if v != data_store[key][k]:
                    print(k,v)
                    print(f"Duplicate entry found for key: {key}. Error.")
                    return
            else:
                data_store[key][k] = v

def remove_vpu(info):
    # 删除字典中的 'vid', 'pid', 'uid' 键
    info.pop('vid', None)
    info.pop('pid', None)
    info.pop('uid', None)
    return info

# 定义一个通用查询函数
def query_info(data_store, vid=None, pid=None, uid=None):
    result = []
    for (v, p, u), info in data_store.items():
        if (vid is None or v == vid) and (pid is None or p == pid) and (uid is None or u == uid):
            result.append({"vid":v, "pid":p, "uid":u})
    return result

def chongfujiancha(data_store):
    vid,pid,uid={},{},{}
    for key, value in data_store.items():
        if key[0] not in vid:
            vid[key[0]] = 0
        else:
            vid[key[0]] += 1
        if key[1] not in pid:
            pid[key[1]] = 0
        else:
            pid[key[1]] += 1
        if key[2] not in uid:
            uid[key[2]] = 0
        else:
            uid[key[2]] += 1
    for key, value in vid.items():
        if value > 1:
            print(f"vid: {key} count: {value}")
    for key, value in pid.items():
        if value > 1:
            print(f"pid: {key} count: {value}")
    for key, value in uid.items():
        if value > 1:
            print(f"uid: {key} count: {value}")
            # query_result = query_info(data_store, uid=key)
            # for single_query in query_result:
            #     print(data_store[single_query.get("vid"), single_query.get("pid"), single_query.get("uid")])

def make_data(data_store, data_type):
    if data_type == "train":
        all_x = ['videos', 'posts', 'users', 'popularity']
    elif data_type == "test":
        all_x = ['videos', 'posts', 'users']

    for file_x in all_x:
        file = f"SMP-Video_anonymized_{file_x}_{data_type}.jsonl"
        with open(os.path.join('./dataset', data_type, file), "r", encoding="utf-8") as f:
            # 逐行读取文件
            for line in f:
                # 将每一行解析为JSON对象
                data = json.loads(line)
                query_result = query_info(data_store, vid=data.get('vid'), pid=data.get('pid'), uid=data.get('uid'))
                if len(query_result)>0:
                    for single_query in query_result:
                        add_info(data_store, vid=single_query.get("vid"), pid=single_query.get("pid"), uid=single_query.get("uid"), info=remove_vpu(data))
                else:
                    add_info(data_store, vid=data.get('vid'), pid=data.get('pid'), uid=data.get('uid'), info=remove_vpu(data))

    data_excel=[]
    excel_name = ['vid', 'pid', 'uid']
    for (v,p,u), info in data_store.items():
        excel_name = list(OrderedDict.fromkeys(excel_name+list(info.keys())))
    for (v,p,u), info in data_store.items():
        data_excel.append([v, p, u] + [info.get(x) for x in excel_name[3:]])
    df = pd.DataFrame(data_excel, columns=excel_name)
    df.to_excel(f"{data_type}.xlsx", index=False, engine="openpyxl")

    # 加载工作簿
    wb = load_workbook(f"{data_type}.xlsx")
    ws = wb.active

    # 设置标题行无边框
    for cell in ws[1]:
        cell.border = Border(left=Side(style=None), 
                            right=Side(style=None), 
                            top=Side(style=None), 
                            bottom=Side(style=None))
        # 设置字体为不加粗
        cell.font = Font(bold=False)
        # 设置水平对齐方式为左对齐
        cell.alignment = Alignment(horizontal="left")

    # 自适应列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # 获取列的字母标识
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存工作簿
    wb.save(f"./dataset/{data_type}.xlsx")

train_data, test_data = {}, {}

make_data(train_data, "train")
make_data(test_data, "test")
